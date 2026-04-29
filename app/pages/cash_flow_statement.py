from nicegui import ui
from app.components.state import state
from app.components.ui_helpers import show_toast, format_amount, navigate
from database_v3 import (
    get_ledgers, get_cash_flow_statement,
)

def _build_pdf(filepath, title, data, headers, row_fn):
    """通用 PDF 导出（使用 reportlab）"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    font_paths = [
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    ]
    font_name = "Helvetica"
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CJK", fp))
                font_name = "CJK"
                break
            except Exception:
                pass
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontName=font_name, fontSize=16)
    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 5*mm))
    table_data = [headers]
    for item in data:
        table_data.append(row_fn(item))
    col_width = (landscape(A4)[0] - 20*mm) / len(headers)
    table = Table(table_data, colWidths=[col_width]*len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1677FF")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(table)
    doc.build(elements)


def _export_cash_flow_excel(method):
    """导出现金流量表为Excel"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"现金流量表_{state.selected_year}{state.selected_month:02d}_{ts}.xlsx"
        fpath = os.path.join(export_dir, fname)
        data = get_cash_flow_statement(lid, state.selected_year, state.selected_month, method)
        if not data:
            show_toast("无现金流量表数据", "warning")
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "现金流量表"
        headers = ["项目", "金额", "备注"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        items = data if isinstance(data, list) else data.get("items", data.get("rows", []))
        for row in items:
            if isinstance(row, dict):
                ws.append([row.get("item", row.get("name", "")), row.get("amount", 0), row.get("note", "")])
            else:
                ws.append([str(row), "", ""])
            for cell in ws[ws.max_row]:
                cell.border = border
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 20
        wb.save(fpath)
        show_toast(f"✅ 现金流量表已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ Excel导出失败: {e}", "error")


def _export_cash_flow_pdf(method):
    """导出现金流量表为PDF"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"现金流量表_{state.selected_year}{state.selected_month:02d}_{ts}.pdf"
        fpath = os.path.join(export_dir, fname)
        data = get_cash_flow_statement(lid, state.selected_year, state.selected_month, method)
        if not data:
            show_toast("无现金流量表数据", "warning")
            return
        items = data if isinstance(data, list) else data.get("items", data.get("rows", []))
        _build_pdf(fpath, "现金流量表", items,
                   ["项目","金额","备注"],
                   lambda r: [r.get("item", r.get("name", "")) if isinstance(r, dict) else str(r),
                              f'¥{r.get("amount", 0):,.2f}' if isinstance(r, dict) and r.get("amount") else "",
                              r.get("note", "") if isinstance(r, dict) else ""])
        show_toast(f"✅ 现金流量表PDF已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ PDF导出失败: {e}", "error")


def render_cash_flow_statement():
    """现金流量表 — 经营活动/投资活动/筹资活动三大类，直接法"""
    if not state.selected_ledger_id:
        ledgers = get_ledgers()
        if ledgers:
            state.selected_ledger_id = ledgers[0]["id"]
    lid = state.selected_ledger_id
    if not lid:
        return

    with ui.card().classes("w-full"):
        with ui.card_section().classes("py-2 px-4 bg-grey-5 border-b border-grey-2"):
            with ui.row().classes("items-center gap-3"):
                ui.label("💧 现金流量表").classes("text-base font-bold")
                ui.separator().props("vertical")
                cf_year_sel = ui.select(options=list(range(2020, 2031)), value=state.selected_year, label="年度").props("dense outlined").classes("w-28")
                cf_month_sel = ui.select(options=list(range(1, 13)), value=state.selected_month, label="月份").props("dense outlined").classes("w-24")
                ui.separator().props("vertical")
                cf_method = ui.toggle(
                    options=[{"label": "直接法", "value": "direct"}, {"label": "间接法", "value": "indirect"}],
                    value="direct"
                ).props("dense").classes("text-xs")
                ui.separator().props("vertical")
                ui.button("📥 导出Excel", color="green",
                          on_click=lambda: _export_cash_flow_excel(cf_method.value)).props("dense").classes("text-xs")
                ui.button("📄 导出PDF", color="blue",
                          on_click=lambda: _export_cash_flow_pdf(cf_method.value)).props("dense").classes("text-xs")

                def _on_cf_period():
                    state.selected_year = cf_year_sel.value
                    state.selected_month = cf_month_sel.value
                    refresh_main()

                cf_year_sel.on("update:value", lambda e: _on_cf_period())
                cf_month_sel.on("update:value", lambda e: _on_cf_period())

    # 获取现金流数据
    try:
        cf = get_cash_flow_statement(lid, state.selected_year, state.selected_month, method=cf_method.value)
    except Exception:
        cf = None

    HC = "text-xs font-semibold uppercase tracking-wide text-grey-6"

    with ui.card().classes("w-full"):
        with ui.card_section().classes("py-2.5 px-4 border-b border-grey-2"):
            ui.label(f"💧 现金流量表 — {state.selected_year}年{state.selected_month}月（{'直接法' if cf_method.value == 'direct' else '间接法'}）").classes("text-base font-bold")

        if not cf:
            with ui.card_section().classes("py-12 text-center"):
                ui.icon("waterfall_chart").style("font-size: 48px; color: var(--gray-300)")
                ui.label("暂无现金流量数据").classes("text-lg font-semibold text-grey-4 mt-4")
            return

        # 构建三大类行数据
        sections = [
            ("operating", "一、经营活动产生的现金流量", "green"),
            ("investing", "二、投资活动产生的现金流量", "blue"),
            ("financing", "三、筹资活动产生的现金流量", "orange"),
        ]

        detail_map = {}
        if cf.get("detail"):
            for d in cf["detail"]:
                detail_map[d.get("type", "")] = d

        rows = []
        for sec_key, sec_label, sec_color in sections:
            sec_data = cf.get(sec_key, {})
            rows.append({
                "name": sec_label, "type": "section_header",
                "inflow": None, "outflow": None, "net": None,
                "color": sec_color, "level": 0, "drillable": False,
            })
            # 流入小计
            inflow = sec_data.get("inflow", 0) or 0
            outflow = sec_data.get("outflow", 0) or 0
            net = sec_data.get("net", 0) or 0
            # 流入明细
            inflow_types = [k for k in detail_map if detail_map[k].get("section") == sec_key and "inflow" in k]
            outflow_types = [k for k in detail_map if detail_map[k].get("section") == sec_key and "outflow" in k]
            for it in inflow_types:
                d = detail_map[it]
                rows.append({
                    "name": f"  {d.get('name', it)}", "type": "detail",
                    "inflow": d.get("net", 0), "outflow": None, "net": None,
                    "color": sec_color, "level": 1, "drillable": True, "cf_type": it,
                })
            if not inflow_types:
                rows.append({
                    "name": "  现金流入小计", "type": "subtotal_row",
                    "inflow": inflow, "outflow": None, "net": None,
                    "color": sec_color, "level": 1, "drillable": False,
                })
            # 流出明细
            for ot in outflow_types:
                d = detail_map[ot]
                rows.append({
                    "name": f"  {d.get('name', ot)}", "type": "detail",
                    "inflow": None, "outflow": abs(d.get("net", 0)), "net": None,
                    "color": sec_color, "level": 1, "drillable": True, "cf_type": ot,
                })
            if not outflow_types:
                rows.append({
                    "name": "  现金流出小计", "type": "subtotal_row",
                    "inflow": None, "outflow": outflow, "net": None,
                    "color": sec_color, "level": 1, "drillable": False,
                })
            # 净额
            rows.append({
                "name": f"  {sec_key == 'operating' and '经营活动现金流量净额' or sec_key == 'investing' and '投资活动现金流量净额' or '筹资活动现金流量净额'}",
                "type": "section_net",
                "inflow": None, "outflow": None, "net": net,
                "color": sec_color, "level": 1, "drillable": False,
            })

        # 现金净增加额
        net_change = cf.get("net_cash_change", 0) or 0
        rows.append({
            "name": "四、现金及现金等价物净增加额", "type": "total",
            "inflow": None, "outflow": None, "net": net_change,
            "color": "grey", "level": 0, "drillable": False,
        })

        cols = [
            {"name": "name", "label": "项 目", "field": "name", "align": "left",
             "headerClasses": HC, "classes": "text-sm text-grey-7", "style": "min-width:220px"},
            {"name": "inflow", "label": "现金流入", "field": "inflow", "align": "right",
             "headerClasses": HC, "classes": "tabular-nums text-sm", "style": "width:140px"},
            {"name": "outflow", "label": "现金流出", "field": "outflow", "align": "right",
             "headerClasses": HC, "classes": "tabular-nums text-sm", "style": "width:140px"},
            {"name": "net", "label": "净 额", "field": "net", "align": "right",
             "headerClasses": HC, "classes": "tabular-nums text-sm", "style": "width:140px"},
        ]

        tbl = ui.table(columns=cols, rows=rows, row_key="name", pagination=False).classes("w-full")

        tbl.add_slot("body-cell-name", r"""
            <q-td key="name" :props="props">
                <span :class="{
                    'font-bold text-base': props.row.type === 'section_header',
                    'font-bold': props.row.type === 'total',
                    'text-green-7': props.row.color === 'green' && props.row.type !== 'detail',
                    'text-blue-7': props.row.color === 'blue' && props.row.type !== 'detail',
                    'text-orange-7': props.row.color === 'orange' && props.row.type !== 'detail',
                    'pl-4': props.row.level === 1,
                    'text-sm text-grey-6': props.row.type === 'detail',
                    'text-sm font-semibold': props.row.type === 'section_net' || props.row.type === 'subtotal_row',
                }"
                :style="props.row.type === 'total' ? 'background:var(--gray-100);' : ''">
                    <span v-if="props.row.drillable" class="cursor-pointer text-blue-7 underline-hover"
                          @click="$parent.$emit('drill', props.row.cf_type)">
                        {{ props.row.name }}
                    </span>
                    <span v-else>{{ props.row.name }}</span>
                </span>
            </q-td>
        """)

        tbl.add_slot("body-cell-inflow", r"""
            <q-td key="inflow" :props="props" class="tabular-nums"
                  :style="['total','section_header'].includes(props.row.type) ? 'background:var(--gray-100);' : ''">
                <span :class="{
                    'font-bold': ['section_header','total','section_net','subtotal_row'].includes(props.row.type),
                    'text-green-7': props.row.inflow > 0,
                }">
                    {{ props.row.inflow !== null && props.row.inflow !== undefined ? '¥' + props.row.inflow.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : '—' }}
                </span>
            </q-td>
        """)

        tbl.add_slot("body-cell-outflow", r"""
            <q-td key="outflow" :props="props" class="tabular-nums"
                  :style="['total','section_header'].includes(props.row.type) ? 'background:var(--gray-100);' : ''">
                <span :class="{
                    'font-bold': ['section_header','total','section_net','subtotal_row'].includes(props.row.type),
                    'text-red-7': props.row.outflow > 0,
                }">
                    {{ props.row.outflow !== null && props.row.outflow !== undefined ? '¥' + props.row.outflow.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : '—' }}
                </span>
            </q-td>
        """)

        tbl.add_slot("body-cell-net", r"""
            <q-td key="net" :props="props" class="tabular-nums"
                  :style="['total','section_header'].includes(props.row.type) ? 'background:var(--gray-100);' : ''">
                <span :class="{
                    'font-bold': ['section_header','total','section_net'].includes(props.row.type),
                    'text-green-7': props.row.net > 0,
                    'text-red-7': props.row.net < 0,
                }">
                    {{ props.row.net !== null && props.row.net !== undefined ? '¥' + props.row.net.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : '—' }}
                </span>
            </q-td>
        """)

        def _on_drill(e):
            cf_type = e.args
            if not cf_type:
                return
            lid = state.selected_ledger_id
            if not lid:
                return
            # 查询该现金流类型下的明细凭证
            rows = query_db("""
                SELECT v.voucher_no, v.date, v.summary,
                       a.code as acct_code, a.name as acct_name,
                       e.debit, e.credit
                FROM vouchers v
                JOIN entries e ON e.voucher_id = v.id
                JOIN accounts a ON a.id = e.account_id
                JOIN entry_cash_flow ecf ON ecf.entry_id = e.id
                JOIN cash_flow_categories cfc ON cfc.id = ecf.cf_category_id
                WHERE v.ledger_id = ? AND cfc.code = ?
                  AND strftime('%Y-%m', v.date) = ?
                ORDER BY v.date DESC
                LIMIT 50
            """, (lid, cf_type, f"{state.selected_year}-{state.selected_month:02d}"))
            if not rows:
                show_toast(f"该现金流类型下无明细凭证", "info")
                return
            # 显示明细弹窗
            with ui.dialog() as dlg, ui.card().classes("w-[600px]"):
                ui.label(f"📋 现金流量明细 — {cf_type}").classes("text-base font-bold mb-2")
                with ui.card_section().classes("max-h-[400px] overflow-y-auto"):
                    cols = [
                        {"name":"voucher_no","label":"凭证号","field":"voucher_no","align":"left"},
                        {"name":"date","label":"日期","field":"date","align":"center"},
                        {"name":"summary","label":"摘要","field":"summary","align":"left"},
                        {"name":"debit","label":"借方","field":"debit","align":"right","classes":"tabular-nums"},
                        {"name":"credit","label":"贷方","field":"credit","align":"right","classes":"tabular-nums"},
                    ]
                    ui.table(columns=cols, rows=rows, row_key="voucher_no",
                             pagination={"rowsPerPage": 10}).classes("w-full text-sm")
                ui.button("关闭", on_click=dlg.close).classes("mt-2")
            dlg.open()

        tbl.on("drill", _on_drill)


