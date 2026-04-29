"""报表 — 科目余额表/资产负债表/利润表"""
import os
from datetime import datetime
from nicegui import ui
from app.components.state import state
from app.components.ui_helpers import show_toast, format_amount, apply_table_style, refresh_main
from database_v3 import (
    get_ledgers, get_account_balances, get_balance_sheet, get_income_statement,
    export_balance_sheet_csv, export_income_statement_csv, export_account_balances_csv,
)

def render_accounts():
    """科目余额表 — ui.table 严格列对齐，金额等宽，借贷分色"""
    if not state.selected_ledger_id:
        ledgers = get_ledgers()
        if ledgers:
            state.selected_ledger_id = ledgers[0]["id"]
    lid = state.selected_ledger_id
    if not lid:
        return
    balances = get_account_balances(lid, state.selected_year, state.selected_month)

    with ui.card().classes("w-full"):
        with ui.card_section().classes("py-2.5 px-4 border-b border-grey-2"):
            ui.label(f"📊 科目余额表 — {state.selected_year}年{state.selected_month}月").classes("text-base font-bold")

        if not balances:
            with ui.card_section().classes("py-12"):
                ui.label("暂无数据").classes("text-sm text-center").style("color:var(--c-text-muted)")
            return

        # 构建行数据
        total = dict.fromkeys(["od","oc","cd","cc","yd","yc","ed","ec"], 0)
        rows = []
        for b in balances:
            od = b.get("opening_dr") or 0; oc = b.get("opening_cr") or 0
            cd = b.get("curr_dr") or 0;  cc = b.get("curr_cr") or 0
            yd = b.get("ytd_dr") or 0;   yc = b.get("ytd_cr") or 0
            ed = b.get("closing_dr") or 0; ec = b.get("closing_cr") or 0
            is_t = b["name"] == "合计"
            if not is_t:
                for k,v in [("od",od),("oc",oc),("cd",cd),("cc",cc),("yd",yd),("yc",yc),("ed",ed),("ec",ec)]:
                    total[k] += v
            rows.append({
                "code": b.get("code",""), "name": b["name"],
                "od": od, "oc": oc, "cd": cd, "cc": cc,
                "yd": yd, "yc": yc, "ed": ed, "ec": ec,
                "level": b.get("level",0), "is_total": is_t,
            })

        # 合计行
        rows.append({
            "code": "", "name": "合 计",
            "od": total["od"], "oc": total["oc"], "cd": total["cd"], "cc": total["cc"],
            "yd": total["yd"], "yc": total["yc"], "ed": total["ed"], "ec": total["ec"],
            "level": 0, "is_total": True,
        })

        HC = "table-header-cell text-uppercase"
        cols = [
            {"name":"code","label":"科目编码","field":"code","align":"left","headerClasses":HC,"classes":"text-xs font-mono","style":"width:96px"},
            {"name":"name","label":"科目名称","field":"name","align":"left","headerClasses":HC,"classes":"text-sm","style":"min-width:140px"},
            {"name":"od","label":"期初借方","field":"od","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
            {"name":"oc","label":"期初贷方","field":"oc","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
            {"name":"cd","label":"本期借方","field":"cd","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
            {"name":"cc","label":"本期贷方","field":"cc","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
            {"name":"yd","label":"累计借方","field":"yd","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
            {"name":"yc","label":"累计贷方","field":"yc","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
            {"name":"ed","label":"期末借方","field":"ed","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
            {"name":"ec","label":"期末贷方","field":"ec","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":"width:112px"},
        ]

        tbl = ui.table(columns=cols, rows=rows, row_key="name", pagination={"rowsPerPage":50}).classes("w-full")

        # 借方列 slot（绿色）
        for col in ["od","cd","yd","ed"]:
            tbl.add_slot(f"body-cell-{col}", r"""
                <q-td :props="props" class="tabular-nums text-sm"
                       :style="props.row.is_total ? 'background:var(--c-bg-hover);font-weight:700;' : (props.row.level===2 ? 'padding-left:24px;' : '')">
                    <span :class="props.row.is_total ? 'font-bold' : 'font-bold'">
                        {{ props.row[""" + col + r"""] ? '¥' + props.row[""" + col + r"""].toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : '—' }}
                    </span>
                </q-td>
            """)

        # 贷方列 slot（红色）
        for col in ["oc","cc","yc","ec"]:
            tbl.add_slot(f"body-cell-{col}", r"""
                <q-td :props="props" class="tabular-nums text-sm"
                       :style="props.row.is_total ? 'background:var(--c-bg-hover);font-weight:700;' : (props.row.level===2 ? 'padding-left:24px;' : '')">
                    <span :class="props.row.is_total ? 'font-bold' : 'font-bold'">
                        {{ props.row[""" + col + r"""] ? '¥' + props.row[""" + col + r"""].toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : '—' }}
                    </span>
                </q-td>
            """)

        tbl.add_slot("body-cell-name", r"""
            <q-td :props="props"
                   :style="props.row.is_total ? 'background:var(--c-bg-hover);font-weight:700;' : ''">
                <span :class="props.row.is_total ? 'font-bold text-base' : (props.row.level === 2 ? 'pl-4 text-sm text-grey-6' : 'text-sm text-secondary')">
                    {{ props.row.name }}
                </span>
            </q-td>
        """)

        # 借贷平衡校验
        with ui.card_section().classes("py-2.5 px-4 bg-grey-50 border-t border-grey-2"):
            with ui.row().classes("justify-center gap-8 items-center text-sm"):
                ui.label("本期借方合计").style("color:var(--c-text-muted)")
                ui.label(f"¥{total['cd']:,.2f}").classes("font-bold tabular-nums").style("color:var(--c-success)")
                ui.label("本期贷方合计").classes("ml-4").style("color:var(--c-text-muted)")
                ui.label(f"¥{total['cc']:,.2f}").classes("font-bold tabular-nums").style("color:var(--c-danger)")
                if abs(total['cd'] - total['cc']) < 0.01:
                    ui.label("✅ 借贷平衡").classes("font-bold ml-4").style("color:var(--c-success)")
                else:
                    ui.label(f"❌ 差额 ¥{abs(total['cd']-total['cc']):,.2f}").classes("font-bold tabular-nums ml-4").style("color:var(--c-danger)")

# ===== 导出功能占位 =====

def _export_balance_sheet():
    """导出资产负债表为Excel"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"资产负债表_{state.selected_year}{state.selected_month:02d}_{ts}.xlsx"
        fpath = os.path.join(export_dir, fname)
        report = get_balance_sheet(lid, state.selected_year, state.selected_month)
        if not report:
            show_toast("无资产负债表数据", "warning")
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "资产负债表"
        headers = ["科目代码", "科目名称", "期初余额", "期末余额", "变动金额", "变动比例"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in report:
            data_row = [row.get("code",""), row.get("name",""),
                       row.get("opening_balance",0) or 0, row.get("balance",0) or 0,
                       row.get("change",0) or 0, row.get("change_pct","")]
            ws.append(data_row)
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        for col, width in zip("ABCDEF", [12, 20, 14, 14, 14, 12]):
            ws.column_dimensions[col].width = width
        wb.save(fpath)
        show_toast(f"✅ 资产负债表已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ Excel导出失败: {e}", "error")

def _export_balance_sheet_pdf():
    """导出资产负债表为PDF"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"资产负债表_{state.selected_year}{state.selected_month:02d}_{ts}.pdf"
        fpath = os.path.join(export_dir, fname)
        report = get_balance_sheet(lid, state.selected_year, state.selected_month)
        if not report:
            show_toast("无资产负债表数据", "warning")
            return
        _build_pdf(fpath, "资产负债表", report,
                   ["科目代码","科目名称","期初余额","期末余额","变动金额","变动比例"],
                   lambda r: [r.get("code",""), r.get("name",""),
                              f'¥{r.get("opening_balance",0) or 0:,.2f}',
                              f'¥{r.get("balance",0) or 0:,.2f}',
                              f'¥{r.get("change",0) or 0:,.2f}',
                              r.get("change_pct","")])
        show_toast(f"✅ 资产负债表PDF已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ PDF导出失败: {e}", "error")

def _export_income_statement():
    """导出利润表为Excel"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"利润表_{state.selected_year}{state.selected_month:02d}_{ts}.xlsx"
        fpath = os.path.join(export_dir, fname)
        report = get_income_statement(lid, state.selected_year, state.selected_month)
        if not report:
            show_toast("无利润表数据", "warning")
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "利润表"
        headers = ["科目代码", "科目名称", "类别", "本期金额", "上期金额", "变动比例"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in report:
            data_row = [row.get("code",""), row.get("name",""), row.get("category",""),
                       row.get("balance",0) or 0, row.get("prev_balance",0) or 0,
                       row.get("change_pct","")]
            ws.append(data_row)
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        for col, width in zip("ABCDEF", [12, 20, 10, 14, 14, 12]):
            ws.column_dimensions[col].width = width
        wb.save(fpath)
        show_toast(f"✅ 利润表已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ Excel导出失败: {e}", "error")


def _build_pdf(filepath, title, data, headers, row_fn):
    """通用 PDF 导出（使用 reportlab）"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    font_paths = [
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    ]
    font_name = "Helvetica"
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CJK", fp))
                font_name = "CJK"
                break
            except Exception:
                pass
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontName=font_name, fontSize=16)
    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 5*mm))
    table_data = [headers]
    for item in data:
        table_data.append(row_fn(item))
    col_width = (landscape(A4)[0] - 20*mm) / len(headers)
    table = Table(table_data, colWidths=[col_width]*len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1677FF")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(table)
    doc.build(elements)


def render_balance_sheet():
    if not state.selected_ledger_id:
        ledgers = get_ledgers()
        if ledgers:
            state.selected_ledger_id = ledgers[0]["id"]
    lid = state.selected_ledger_id
    if not lid:
        return
    """资产负债表 — 左右两栏 ui.table，严格列对齐，同比/环比增强"""
    lid = state.selected_ledger_id

    # ── 期间选择器 ──
    with ui.card().classes("w-full"):
        with ui.card_section().classes("py-2 px-4 border-b border-grey-2").style("color:var(--c-bg-hover)"):
            with ui.row().classes("items-center gap-3"):
                ui.label("📗 资产负债表").classes("text-base font-bold")
                ui.separator().props("vertical")
                year_sel = ui.select(
                    options=list(range(2020, 2031)),
                    value=state.selected_year,
                    label="年度"
                ).props("dense outlined").classes("w-28")
                month_sel = ui.select(
                    options=list(range(1, 13)),
                    value=state.selected_month,
                    label="月份"
                ).props("dense outlined").classes("w-24")
                ui.separator().props("vertical")
                ui.button("📥 导出Excel", color="green", on_click=lambda: _export_balance_sheet())                     .props("dense").classes("text-xs")
                ui.button("📄 导出PDF", color="blue", on_click=lambda: _export_balance_sheet_pdf())                     .props("dense").classes("text-xs")
                ui.separator().props("vertical")
                # 同比/环比切换
                compare_mode = ui.toggle(
                    options={"mom": "环比", "yoy": "同比"},
                    value="mom"
                ).props("dense").classes("text-xs")

                def _on_period_change():
                    state.selected_year = year_sel.value
                    state.selected_month = month_sel.value
                    refresh_main()

                year_sel.on("update:value", lambda e: _on_period_change())
                month_sel.on("update:value", lambda e: _on_period_change())

    bs = get_balance_sheet(lid, state.selected_year, state.selected_month)
    # 获取对比期间数据
    if compare_mode.value == "mom":
        prev_month = state.selected_month - 1
        prev_year = state.selected_year
        if prev_month < 1:
            prev_month = 12
            prev_year -= 1
        label_prev = f"{prev_year}年{prev_month}月"
    else:
        prev_year = state.selected_year - 1
        prev_month = state.selected_month
        label_prev = f"{prev_year}年{prev_month}月"
    bs_prev = get_balance_sheet(lid, prev_year, prev_month)

    HC = "table-header-cell text-uppercase"
    num_style = "width:120px"

    with ui.card().classes("w-full"):
        with ui.card_section().classes("py-2.5 px-4 border-b border-grey-2"):
            ui.label(f"📗 资产负债表 — {bs['date']}").classes("text-base font-bold")

        with ui.row().classes("w-full gap-0"):
            # 左栏：资产
            with ui.column().classes("w-1/2 pr-4"):
                with ui.card_section().classes("py-2 px-3 bg-green-50"):
                    ui.label("资 产").classes("text-sm font-bold text-center uppercase tracking-widest").style("color:var(--c-success)")
                prev_assets = {r["name"]: r.get("end", 0) for r in bs_prev["assets"]}
                rows_a = []
                for r in bs["assets"]:
                    prev_val = prev_assets.get(r["name"], 0)
                    end_val = r.get("end", 0)
                    change_pct = ((end_val - prev_val) / abs(prev_val) * 100) if prev_val else None
                    rows_a.append({
                        "name": r["name"], "code": r.get("code", "") or "",
                        "end": end_val, "open": r.get("open", 0),
                        "prev": prev_val, "change": change_pct,
                        "level": r.get("level", 0)
                    })
                cols_a = [
                    {"name":"name","label":"项 目","field":"name","align":"left","headerClasses":HC,"classes":"text-sm","style":"min-width:120px"},
                    {"name":"code","label":"行次","field":"code","align":"center","headerClasses":HC,"classes":"text-xs font-mono","style":"width:48px"},
                    {"name":"end","label":"期末数","field":"end","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":num_style},
                    {"name":"open","label":"年初数","field":"open","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":num_style},
                    {"name":"prev","label":f"对比({label_prev})","field":"prev","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":num_style},
                    {"name":"change","label":"变化率","field":"change","align":"right","headerClasses":HC,"classes":"tabular-nums text-xs","style":"width:72px"},
                ]
                tbl_a = ui.table(columns=cols_a, rows=rows_a, row_key="name", pagination=False).classes("w-full")
                tbl_a.add_slot("body-cell-name", r"""
                    <q-td :props="props"
                           :style="props.row.level===0 ? 'background:var(--c-success-light);font-weight:700;' : (props.row.level===2 ? 'padding-left:20px;' : '')">
                        <span :class="props.row.level===0 ? 'font-bold text-success' : (props.row.level===2 ? 'text-sm text-secondary' : 'text-sm text-secondary')">
                            {{ props.row.name }}
                        </span>
                    </q-td>
                """)
                for col in ["end","open","prev"]:
                    tbl_a.add_slot(f"body-cell-{col}", r"""
                        <q-td :props="props" class="tabular-nums text-sm"
                               :style="props.row.level===0 ? 'background:var(--c-success-light);font-weight:700;' : ''">
                            <span :class="props.row.level===0 ? 'font-bold text-success' : 'text-primary'">
                                {{ props.row[""" + col + r"""] !== null ? '¥' + props.row[""" + col + r"""].toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : '—' }}
                            </span>
                        </q-td>
                    """)
                tbl_a.add_slot("body-cell-change", r"""
                    <q-td :props="props" class="tabular-nums text-xs"
                           :style="props.row.level===0 ? 'background:var(--c-success-light);font-weight:700;' : ''">
                        <span v-if="props.row.change !== null"
                              :class="props.row.change > 0 ? 'num-positive' : (props.row.change < 0 ? 'num-negative' : 'text-muted')">
                            {{ props.row.change > 0 ? '▲' : (props.row.change < 0 ? '▼' : '—') }}
                            {{ Math.abs(props.row.change).toFixed(1) }}%
                        </span>
                        <span v-else class="text-muted">—</span>
                    </q-td>
                """)

            # 右栏：负债+权益
            with ui.column().classes("w-1/2 pl-4"):
                with ui.card_section().classes("py-2 px-3 bg-blue-50"):
                    ui.label("负债和所有者权益").classes("text-sm font-bold text-center uppercase tracking-widest").style("color:var(--c-primary)")
                prev_le = {r["name"]: r.get("end", 0) for r in bs_prev["liabilities"] + bs_prev["equity"]}
                rows_l = []
                for r in bs["liabilities"] + bs["equity"]:
                    prev_val = prev_le.get(r["name"], 0)
                    end_val = r.get("end", 0)
                    change_pct = ((end_val - prev_val) / abs(prev_val) * 100) if prev_val else None
                    rows_l.append({
                        "name": r["name"], "code": r.get("code", "") or "",
                        "end": end_val, "open": r.get("open", 0),
                        "prev": prev_val, "change": change_pct,
                        "level": r.get("level", 0)
                    })
                cols_l = [
                    {"name":"name","label":"项 目","field":"name","align":"left","headerClasses":HC,"classes":"text-sm","style":"min-width:120px"},
                    {"name":"code","label":"行次","field":"code","align":"center","headerClasses":HC,"classes":"text-xs font-mono","style":"width:48px"},
                    {"name":"end","label":"期末数","field":"end","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":num_style},
                    {"name":"open","label":"年初数","field":"open","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":num_style},
                    {"name":"prev","label":f"对比({label_prev})","field":"prev","align":"right","headerClasses":HC,"classes":"tabular-nums text-sm","style":num_style},
                    {"name":"change","label":"变化率","field":"change","align":"right","headerClasses":HC,"classes":"tabular-nums text-xs","style":"width:72px"},
                ]
                tbl_l = ui.table(columns=cols_l, rows=rows_l, row_key="name", pagination=False).classes("w-full")
                tbl_l.add_slot("body-cell-name", r"""
                    <q-td :props="props"
                           :style="props.row.level===0 ? 'background:var(--c-primary-light);font-weight:700;' : (props.row.level===2 ? 'padding-left:20px;' : '')">
                        <span :class="props.row.level===0 ? 'font-bold text-primary' : (props.row.level===2 ? 'text-sm text-secondary' : 'text-sm text-secondary')">
                            {{ props.row.name }}
                        </span>
                    </q-td>
                """)
                for col in ["end","open","prev"]:
                    tbl_l.add_slot(f"body-cell-{col}", r"""
                        <q-td :props="props" class="tabular-nums text-sm"
                               :style="props.row.level===0 ? 'background:var(--c-primary-light);font-weight:700;' : ''">
                            <span :class="props.row.level===0 ? 'font-bold text-primary' : 'text-primary'">
                                {{ props.row[""" + col + r"""] !== null ? '¥' + props.row[""" + col + r"""].toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : '—' }}
                            </span>
                        </q-td>
                    """)
                tbl_l.add_slot("body-cell-change", r"""
                    <q-td :props="props" class="tabular-nums text-xs"
                           :style="props.row.level===0 ? 'background:var(--c-primary-light);font-weight:700;' : ''">
                        <span v-if="props.row.change !== null"
                              :class="props.row.change > 0 ? 'num-positive' : (props.row.change < 0 ? 'num-negative' : 'text-muted')">
                            {{ props.row.change > 0 ? '▲' : (props.row.change < 0 ? '▼' : '—') }}
                            {{ Math.abs(props.row.change).toFixed(1) }}%
                        </span>
                        <span v-else class="text-muted">—</span>
                    </q-td>
                """)

        # 平衡校验
        with ui.card_section().classes("py-3 px-4 bg-grey-50 border-t border-grey-2"):
            ta, tl, te = bs["total_assets"], bs["total_liab"], bs["total_equity"]
            diff = abs(ta - (tl + te))
            with ui.row().classes("justify-center gap-4 items-center text-sm"):
                for lbl, val, clr in [("资产总计",ta,"text-success"),("负债合计",tl,"text-danger"),("所有者权益",te,"text-primary")]:
                    ui.label(lbl).style("color:var(--c-text-muted)")
                    ui.label(f"¥{val:,.2f}").classes(f"{clr} font-bold tabular-nums text-base")
                if diff < 0.01:
                    ui.label("✅ 平衡").classes("font-bold ml-2").style("color:var(--c-success)")
                else:
                    ui.label(f"❌ 差额 ¥{diff:,.2f}").classes("font-bold tabular-nums ml-2").style("color:var(--c-danger)")
def render_income_statement():
    if not state.selected_ledger_id:
        ledgers = get_ledgers()
        if ledgers:
            state.selected_ledger_id = ledgers[0]["id"]
    lid = state.selected_ledger_id
    if not lid:
        return
    """利润表 — 项目 | 行次 | 本年累计 | 本月金额 | 同比变化"""
    lid = state.selected_ledger_id
    inc = get_income_statement(lid, state.selected_year, state.selected_month)
    # 去年同期数据
    inc_yoy = get_income_statement(lid, state.selected_year - 1, state.selected_month)

    with ui.card().classes("w-full"):
        with ui.card_section().classes("py-2 px-4 border-b border-grey-2").style("color:var(--c-bg-hover)"):
            with ui.row().classes("items-center gap-3"):
                ui.label("📈 利润表").classes("text-base font-bold")
                ui.separator().props("vertical")
                inc_year_sel = ui.select(options=list(range(2020,2031)), value=state.selected_year, label="年度").props("dense outlined").classes("w-28")
                inc_month_sel = ui.select(options=list(range(1,13)), value=state.selected_month, label="月份").props("dense outlined").classes("w-24")
                ui.separator().props("vertical")
                ui.button("📥 导出Excel", color="green", on_click=lambda: _export_income_statement()).props("dense").classes("text-xs")
                def _on_inc_period():
                    state.selected_year = inc_year_sel.value
                    state.selected_month = inc_month_sel.value
                    refresh_main()
                inc_year_sel.on("update:value", lambda e: _on_inc_period())
                inc_month_sel.on("update:value", lambda e: _on_inc_period())
        with ui.card_section().classes("py-1.5 px-4 border-b").style("border-color:var(--c-border);background:var(--c-bg-hover)"):
            ui.label(f"📈 利润表 — {inc['date']}").classes("text-sm font-semibold").style("color:var(--c-text-secondary)")

        cols = [
            {"name":"name","label":"项 目","field":"name","align":"left","headerClasses":"table-header-cell text-uppercase"},
            {"name":"code","label":"行次","field":"code","align":"center","headerClasses":"table-header-cell text-uppercase"},
            {"name":"ytd","label":"本年累计","field":"ytd","align":"right","classes":"tabular-nums text-sm","headerClasses":"table-header-cell text-uppercase"},
            {"name":"month","label":"本月金额","field":"month","align":"right","classes":"tabular-nums text-sm","headerClasses":"table-header-cell text-uppercase"},
            {"name":"yoy","label":"去年同期","field":"yoy","align":"right","classes":"tabular-nums text-sm","headerClasses":"table-header-cell text-uppercase"},
            {"name":"yoy_pct","label":"同比%","field":"yoy_pct","align":"right","classes":"tabular-nums text-xs","headerClasses":"table-header-cell text-uppercase","style":"width:72px"},
        ]
        yoy_map = {r["name"]: r.get("ytd") for r in inc_yoy["rows"]}
        rows = []
        for r in inc["rows"]:
            ytd_val = r.get("ytd")
            yoy_val = yoy_map.get(r["name"])
            yoy_pct = ((ytd_val - yoy_val) / abs(yoy_val) * 100) if (yoy_val and ytd_val is not None) else None
            rows.append({
                "name": r["name"], "code": r.get("code", "") or "",
                "ytd": ytd_val, "month": r.get("month"),
                "yoy": yoy_val, "yoy_pct": yoy_pct,
                "type": r.get("type", ""), "level": r.get("level", 0)
            })

        tbl = ui.table(columns=cols, rows=rows, row_key="name", pagination=False).classes("w-full")

        BOLD_TYPES = "header rev_total expense_header revenue_header subtotal total"
        tbl.add_slot("body-cell-name", r"""
            <q-td key="name" :props="props">
                <span :class="{
                    'font-bold': ['header','rev_total','expense_header','revenue_header','subtotal','total'].includes(props.row.type),
                    'font-bold': props.row.type === 'expense_header',
                    'font-bold': props.row.type === 'revenue_header',
                    'pl-4': props.row.level === 1,
                    'pl-8': props.row.level === 2,
                }" :style="['subtotal','total','rev_total'].includes(props.row.type) ? 'background:var(--c-bg-hover);' : ''">
                    {{ props.row.name }}
                </span>
            </q-td>
        """)
        tbl.add_slot("body-cell-ytd", r"""
            <q-td key="ytd" :props="props" class="tabular-nums"
                  :style="['subtotal','total','rev_total'].includes(props.row.type) ? 'background:var(--c-bg-hover);' : ''">
                <span :class="{
                    'font-bold': ['header','rev_total','expense_header','revenue_header','subtotal','total'].includes(props.row.type),
                    'font-bold': props.row.ytd < 0,
                }">
                    {{ props.row.ytd !== null ? '¥' + (props.row.ytd < 0 ? '(' + Math.abs(props.row.ytd).toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) + ')' : props.row.ytd.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2})) : '—' }}
                </span>
            </q-td>
        """)
        tbl.add_slot("body-cell-month", r"""
            <q-td key="month" :props="props" class="tabular-nums"
                  :style="['subtotal','total','rev_total'].includes(props.row.type) ? 'background:var(--c-bg-hover);' : ''">
                <span :class="{
                    'font-bold': ['header','rev_total','expense_header','revenue_header','subtotal','total'].includes(props.row.type),
                    'font-bold': props.row.month < 0,
                }">
                    {{ props.row.month !== null ? '¥' + (props.row.month < 0 ? '(' + Math.abs(props.row.month).toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) + ')' : props.row.month.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2})) : '—' }}
                </span>
            </q-td>
        """)
        tbl.add_slot("body-cell-code", r"""
            <q-td key="code" :props="props" class="text-muted text-xs font-mono tabular-nums">
                {{ props.row.code }}
            </q-td>
        """)

