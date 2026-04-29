"""导入导出"""
import os, tempfile
from datetime import datetime
from nicegui import ui
from app.components.state import state
from app.components.ui_helpers import show_toast, refresh_main
from database_v3 import (
    get_ledgers, import_vouchers_from_excel, export_vouchers_csv,
    export_balance_sheet_csv, export_income_statement_csv, export_account_balances_csv,
)

def render_import():
    if not state.selected_ledger_id:
        ledgers = get_ledgers()
        if ledgers:
            state.selected_ledger_id = ledgers[0]["id"]
    lid = state.selected_ledger_id
    import tempfile, os

    with ui.row().classes("w-full gap-3"):
        # 左侧：上传 + 模板下载
        with ui.column().classes("w-1/2 gap-2"):
            with ui.card().classes("w-full"):
                with ui.card_section().classes("py-2 px-3"):
                    ui.label("📥 上传 Excel 凭证文件").classes("text-sm font-semibold")
                with ui.card_section().classes("py-1 px-3"):
                    ui.label("Excel 格式：日期 | 摘要 | 科目代码 | 科目名称 | 借方 | 贷方").classes("text-sm").style("color:var(--c-text-muted)")
                    ui.label("同一凭证的行保持相同日期+摘要，程序自动合并").classes("text-sm").style("color:var(--c-text-muted)")

                    upload = ui.upload(
                        on_upload=lambda e: do_import(e),
                        auto_upload=True,
                        label="点击或拖拽上传 .xlsx 文件",
                        max_file_size=10*1024*1024,
                    ).props("accept=.xlsx,.xls").classes("w-full mt-2")

            with ui.card().classes("w-full"):
                with ui.card_section().classes("py-2 px-3"):
                    ui.label("📋 Excel 模板").classes("text-sm font-semibold")
                with ui.card_section().classes("py-1 px-3"):
                    ui.button("⬇️ 下载模板文件", color="blue",
                              on_click=lambda: ui.download(_generate_template())).props("dense")

        # 右侧：导入结果
        with ui.column().classes("w-1/2"):
            with ui.card().classes("w-full"):
                with ui.card_section().classes("py-2 px-3"):
                    ui.label("📋 导入说明").classes("text-sm font-semibold")
                with ui.card_section().classes("py-1 px-3"):
                    with ui.column().classes("gap-1 text-sm").style("color:var(--c-text-secondary)"):
                        tips = [
                            "Excel 第一行为表头，从第二行开始填数据",
                            "同一凭证的多行分录保持相同「日期」和「摘要」",
                            "系统自动按日期+摘要合并生成分录",
                            "借贷不平的凭证会被跳过并提示",
                            "导入的凭证默认为草稿状态，请手动过账",
                        ]
                        for i, tip in enumerate(tips, 1):
                            with ui.row().classes("items-start gap-2"):
                                ui.label(f"{i}.").classes("font-mono text-xs w-4 text-right").style("color:var(--c-text-muted)")
                                ui.label(tip).classes("flex-grow text-sm")


def do_import(upload_event):
    import tempfile, os
    lid = state.selected_ledger_id
    try:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(upload_event.content.read())
        tmp.close()
        result = import_vouchers_from_excel(lid, tmp.name)
        os.unlink(tmp.name)

        # 增强结果展示：部分成功 + 错误详情弹窗
        imported = result.get("imported", 0)
        errors = result.get("errors", [])

        if imported > 0:
            msg = f"✅ 成功导入 {imported} 张凭证"
            if errors:
                msg += f"（{len(errors)} 行有误）"
            show_toast(msg, "success" if not errors else "warning")
            if errors:
                with ui.dialog() as err_dlg, ui.card().classes("w-[500px]"):
                    ui.label("⚠️ 导入错误详情").classes("text-base font-bold mb-2")
                    with ui.card_section().classes("max-h-[300px] overflow-y-auto"):
                        for err in errors[:20]:
                            ui.label(str(err)).classes("text-sm").style("color:var(--c-danger)")
                    ui.button("关闭", on_click=err_dlg.close).classes("mt-2")
                err_dlg.open()
            refresh_main()
        elif errors:
            show_toast(f"导入失败：{errors[0]}", "error")
        else:
            show_toast("没有可导入的数据", "warning")
    except Exception as e:
        show_toast(f"❌ 导入失败: {e}", "error")


def _generate_template():
    """生成 Excel 模板文件"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile, os

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "凭证导入模板"

    headers = ["日期", "摘要", "科目代码", "科目名称", "借方金额", "贷方金额"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 示例数据
    examples = [
        ["2026-04-01", "收到投资款", "1002", "银行存款", 1000000, 0],
        ["2026-04-01", "收到投资款", "4001", "实收资本", 0, 1000000],
        ["2026-04-05", "购买办公用品", "6602", "管理费用", 5000, 0],
        ["2026-04-05", "购买办公用品", "1002", "银行存款", 0, 5000],
    ]
    for row in examples:
        ws.append(row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name


# ===== 5. 期末结转 =====

def _export_vouchers_excel():
    """导出凭证列表为 Excel"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"凭证列表_{state.selected_year}{state.selected_month:02d}_{ts}.xlsx"
        fpath = os.path.join(export_dir, fname)
        from database_v3 import get_vouchers
        vouchers = get_vouchers(lid, state.selected_year, state.selected_month)
        if not vouchers:
            show_toast("当月无凭证数据", "warning")
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "凭证列表"
        headers = ["凭证号", "日期", "摘要", "科目代码", "科目名称", "借方", "贷方", "状态"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for v in vouchers:
            row = [v.get("voucher_no",""), v.get("date",""), v.get("summary",""),
                   v.get("account_code",""), v.get("account_name",""),
                   v.get("debit",0) or 0, v.get("credit",0) or 0, v.get("status","")]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        for col, width in zip("ABCDEFGH", [14, 12, 24, 12, 16, 14, 14, 10]):
            ws.column_dimensions[col].width = width
        wb.save(fpath)
        show_toast(f"✅ 凭证列表已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ Excel导出失败: {e}", "error")

def _export_vouchers_pdf():
    """导出凭证列表为 PDF"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"凭证列表_{state.selected_year}{state.selected_month:02d}_{ts}.pdf"
        fpath = os.path.join(export_dir, fname)
        from database_v3 import get_vouchers
        vouchers = get_vouchers(lid, state.selected_year, state.selected_month)
        if not vouchers:
            show_toast("当月无凭证数据", "warning")
            return
        _build_pdf(fpath, "凭证列表", vouchers,
                   ["凭证号","日期","摘要","科目代码","科目名称","借方","贷方","状态"],
                   lambda v: [v.get("voucher_no",""), v.get("date",""), v.get("summary",""),
                              v.get("account_code",""), v.get("account_name",""),
                              f'¥{v.get("debit",0) or 0:,.2f}', f'¥{v.get("credit",0) or 0:,.2f}', v.get("status","")])
        show_toast(f"✅ 凭证列表PDF已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ PDF导出失败: {e}", "error")

def _export_account_balances_excel():
    """导出科目余额表为 Excel"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"科目余额表_{state.selected_year}{state.selected_month:02d}_{ts}.xlsx"
        fpath = os.path.join(export_dir, fname)
        from database_v3 import get_account_balances
        balances = get_account_balances(lid, state.selected_year, state.selected_month)
        if not balances:
            show_toast("无科目余额数据", "warning")
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "科目余额表"
        headers = ["科目代码", "科目名称", "类别", "期初余额", "本期借方", "本期贷方", "期末余额"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for b in balances:
            row = [b.get("code",""), b.get("name",""), b.get("category",""),
                   b.get("opening_balance",0) or 0, b.get("debit",0) or 0,
                   b.get("credit",0) or 0, b.get("balance",0) or 0]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        for col, width in zip("ABCDEFG", [12, 16, 10, 14, 14, 14, 14]):
            ws.column_dimensions[col].width = width
        wb.save(fpath)
        show_toast(f"✅ 科目余额表已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ Excel导出失败: {e}", "error")

def _export_account_balances_pdf():
    """导出科目余额表为 PDF"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"科目余额表_{state.selected_year}{state.selected_month:02d}_{ts}.pdf"
        fpath = os.path.join(export_dir, fname)
        from database_v3 import get_account_balances
        balances = get_account_balances(lid, state.selected_year, state.selected_month)
        if not balances:
            show_toast("无科目余额数据", "warning")
            return
        _build_pdf(fpath, "科目余额表", balances,
                   ["科目代码","科目名称","类别","期初余额","本期借方","本期贷方","期末余额"],
                   lambda b: [b.get("code",""), b.get("name",""), b.get("category",""),
                              f'¥{b.get("opening_balance",0) or 0:,.2f}',
                              f'¥{b.get("debit",0) or 0:,.2f}',
                              f'¥{b.get("credit",0) or 0:,.2f}',
                              f'¥{b.get("balance",0) or 0:,.2f}'])
        show_toast(f"✅ 科目余额表PDF已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ PDF导出失败: {e}", "error")

def _export_journal_excel():
    """导出凭证分录为 Excel"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"凭证分录_{state.selected_year}{state.selected_month:02d}_{ts}.xlsx"
        fpath = os.path.join(export_dir, fname)
        from database_v3 import get_vouchers
        vouchers = get_vouchers(lid, state.selected_year, state.selected_month)
        if not vouchers:
            show_toast("当月无凭证数据", "warning")
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "凭证分录"
        headers = ["凭证号", "日期", "摘要", "科目代码", "科目名称", "借方金额", "贷方金额"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="52C41A", end_color="52C41A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for v in vouchers:
            row = [v.get("voucher_no",""), v.get("date",""), v.get("summary",""),
                   v.get("account_code",""), v.get("account_name",""),
                   v.get("debit",0) or 0, v.get("credit",0) or 0]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        for col, width in zip("ABCDEFG", [14, 12, 24, 12, 16, 14, 14]):
            ws.column_dimensions[col].width = width
        wb.save(fpath)
        show_toast(f"✅ 凭证分录已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ Excel导出失败: {e}", "error")

def _export_journal_pdf():
    """导出凭证分录为 PDF"""
    try:
        lid = state.selected_ledger_id
        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"凭证分录_{state.selected_year}{state.selected_month:02d}_{ts}.pdf"
        fpath = os.path.join(export_dir, fname)
        from database_v3 import get_vouchers
        vouchers = get_vouchers(lid, state.selected_year, state.selected_month)
        if not vouchers:
            show_toast("当月无凭证数据", "warning")
            return
        _build_pdf(fpath, "凭证分录", vouchers,
                   ["凭证号","日期","摘要","科目代码","科目名称","借方金额","贷方金额"],
                   lambda v: [v.get("voucher_no",""), v.get("date",""), v.get("summary",""),
                              v.get("account_code",""), v.get("account_name",""),
                              f'¥{v.get("debit",0) or 0:,.2f}', f'¥{v.get("credit",0) or 0:,.2f}'])
        show_toast(f"✅ 凭证分录PDF已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ PDF导出失败: {e}", "error")

def _build_pdf(filepath, title, data, headers, row_fn):
    """通用 PDF 导出（使用 reportlab）"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    # 注册中文字体
    font_paths = [
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    ]
    font_name = "Helvetica"
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CJK", fp))
                font_name = "CJK"
                break
            except Exception:
                pass
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontName=font_name, fontSize=16)
    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 5*mm))
    table_data = [headers]
    for item in data:
        table_data.append(row_fn(item))
    col_width = (landscape(A4)[0] - 20*mm) / len(headers)
    table = Table(table_data, colWidths=[col_width]*len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1677FF")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
def render_export():
    if not state.selected_ledger_id:
        ledgers = get_ledgers()
        if ledgers:
            state.selected_ledger_id = ledgers[0]["id"]
    lid = state.selected_ledger_id
    if not lid:
        return
    lid = state.selected_ledger_id
    export_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(export_dir, exist_ok=True)

    with ui.card().classes("w-full"):
        with ui.card_section().classes("py-2.5 px-4 border-b border-grey-2"):
            ui.label("📦 数据导出").classes("text-base font-bold")

        with ui.card_section().classes("py-3 px-4"):
            with ui.row().classes("gap-3"):
                for etype, title, desc, color in [
                    ("vouchers","📋 凭证列表","导出当月所有凭证","blue"),
                    ("balances","📊 科目余额表","导出科目余额表","green"),
                    ("balance_sheet","📗 资产负债表","导出资产负债表","purple"),
                    ("income_statement","📈 利润表","导出利润表","orange"),
                ]:
                    with ui.card().classes("flex-1"):
                        with ui.card_section().classes("py-2 px-3"):
                            ui.label(title).classes("text-sm font-semibold")
                            ui.label(desc).classes("text-xs mt-0.5").style("color:var(--c-text-muted)")
                        with ui.card_section().classes("py-1.5 px-3"):
                            ui.button("导出 CSV", color=color, on_click=lambda t=etype: do_export(t)).props("dense").classes("w-full")

        export_files = sorted(os.listdir(export_dir), reverse=True)[:10] if os.path.exists(export_dir) else []
        if export_files:
            with ui.card_section().classes("py-2 px-4 border-t border-grey-1"):
                ui.label("最近导出").classes("text-xs font-semibold uppercase tracking-wide mb-1").style("color:var(--c-text-muted)")
                for fname in export_files:
                    fpath = os.path.join(export_dir, fname)
                    size = os.path.getsize(fpath)
                    sz = f"{size/1024:.1f} KB" if size > 1024 else f"{size} B"
                    with ui.card_section().classes("py-1 px-4"):
                        with ui.row().classes("items-center gap-2 w-full"):
                            ui.icon("description", size="sm").style("color:var(--c-text-muted)")
                            ui.label(fname).classes("flex-grow text-sm truncate")
                            ui.label(sz).classes("text-xs tabular-nums").style("color:var(--c-text-muted)")
                            ui.button("下载", color="blue", on_click=lambda fp=fpath: ui.download(fp)).props("dense flat text-xs")
        else:
            with ui.card_section().classes("py-6 px-4"):
                ui.label("暂无导出文件").classes("text-sm text-center").style("color:var(--c-text-muted)")
def do_export(export_type):
    lid = state.selected_ledger_id
    export_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(export_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        if export_type == "vouchers":
            fname = f"凭证列表_{state.selected_year}{state.selected_month:02d}_{ts}.csv"
            fpath = os.path.join(export_dir, fname)
            count = export_vouchers_csv(lid, state.selected_year, state.selected_month, fpath)
            show_toast(f"✅ 已导出 {count} 条凭证 → {fname}", "success")
        elif export_type == "balances":
            fname = f"科目余额表_{state.selected_year}{state.selected_month:02d}_{ts}.csv"
            fpath = os.path.join(export_dir, fname)
            count = export_account_balances_csv(lid, state.selected_year, state.selected_month, fpath)
            show_toast(f"✅ 已导出 {count} 个科目 → {fname}", "success")
        elif export_type == "bs":
            fname = f"资产负债表_{state.selected_year}{state.selected_month:02d}_{ts}.csv"
            fpath = os.path.join(export_dir, fname)
            export_balance_sheet_csv(lid, state.selected_year, state.selected_month, fpath)
            show_toast(f"✅ 资产负债表已导出 → {fname}", "success")
        elif export_type == "is":
            fname = f"利润表_{state.selected_year}{state.selected_month:02d}_{ts}.csv"
            fpath = os.path.join(export_dir, fname)
            export_income_statement_csv(lid, state.selected_year, state.selected_month, fpath)
            show_toast(f"✅ 利润表已导出 → {fname}", "success")
        refresh_main()
    except Exception as e:
        show_toast(f"❌ 导出失败: {e}", "error")


# ===== 7. AI 助手 =====

